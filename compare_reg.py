from __future__ import print_function
import sys, os, xml.etree.ElementTree as ET
from filecmp import dircmp
from openpyxl import load_workbook, worksheet
import xlrd
from sets import Set

def xrld(file):
	filename, ext=os.path.splitext(file)
	
	dirname = os.path.join(os.path.dirname(file),"_temp")
	try:
		os.mkdir(dirname)
	except Exception as e:
		pass
		
	basename=os.path.basename(file)
	tmpfile=os.path.join(dirname, os.path.splitext(basename)[0])
	
	files=[]
	
	tree=ET.parse(file)
	root=tree.getroot()
	for ws in root.findall("{urn:schemas-microsoft-com:office:spreadsheet}Worksheet"):
		sn=ws.attrib["{urn:schemas-microsoft-com:office:spreadsheet}Name"]
		with open("".join([tmpfile, "_", sn, ".txt"]), "w") as f:
			files.append(os.path.basename(f.name))
			for row in ws.findall(".//{urn:schemas-microsoft-com:office:spreadsheet}Row"):
				values=[]
				for cell in row.findall(".//{urn:schemas-microsoft-com:office:spreadsheet}Data"):
					if cell.text is None:
						values.append("")
					else:
						values.append(cell.text)

				f.write(",".join(values))
				f.write("\n")

	return [basename, files]
			
def openpyxl(file):
    #read workbook to get data
    filename, ext=os.path.splitext(file)
	
    dirname = os.path.join(os.path.dirname(file),"_temp")
    try:
		os.mkdir(dirname)
    except Exception as e:
		pass
		
    basename=os.path.basename(file)
    tmpfile=os.path.join(dirname, os.path.splitext(basename)[0])

    wb = load_workbook(filename=file)
	
    files=[]
	
    for sn in wb.get_sheet_names():
		#Iterate through worksheet and print cell contents
	    with open("".join([tmpfile, "_", sn, ".txt"]), "w") as f:
			files.append(os.path.basename(f.name))
			ws=wb.get_sheet_by_name(sn)
			for row in ws.iter_rows():
				values=[]
				for cell in row:
					if cell.value is None:
						values.append("")
					elif isinstance(cell.value, unicode):
						values.append(cell.value.encode('utf8'))
					else:
						values.append(str(cell.value))
			
				f.write(",".join(values))
				f.write("\n")

    return [basename, files]

def generateTextFile(folder, mapping=[]):
	for root, dirs, files in os.walk(folder):
		files = [file for file in files if os.path.splitext(file)[1].endswith('xls') or os.path.splitext(file)[1].endswith('xlsx')]
		for file in files:
			fn, ext = os.path.splitext(file)
			if ext == ".xls":
				mapping.append(xrld(os.path.join(root, file)))
			elif ext == ".xlsx":
				mapping.append(openpyxl(os.path.join(root, file)))
			else:
				print(file)
	return mapping;

def collectDiffFiles(dcmp, files={"differences":[], "baseline missing":[], "post missing":[]}):
	for name in (dcmp.diff_files):
		if not name.endswith("xls") and not name.endswith("xlsx"):
			files["differences"].append([dcmp.left, dcmp.right, name])
	for name in (dcmp.left_only):
		files["post missing"].append(name)
	for name in (dcmp.right_only):
		files["baseline missing"].append(name)
	for sub_dcmp in dcmp.subdirs.values():
		collectDiffFiles(sub_dcmp)
	return files;
				
def sortAndCompare(files):
	trueDiffs=[]
	for file in files:
		with open(os.path.join(file[0], file[2])) as bf:
			baseline = sorted(bf.readlines())
		with open(os.path.join(file[1], file[2])) as pf:
			post = sorted(pf.readlines())
		if baseline != post:
			trueDiffs.append(file[2])
	return trueDiffs

def main():
	if len(sys.argv) != 3:
		print("Usage: python compare_reg.py <baseline folder to be processed> <post folder to be processed>")
		sys.exit(1)
        sys.argv[1] = os.path.normpath(sys.argv[1])
        sys.argv[2] = os.path.normpath(sys.argv[2])
        if not os.path.exists(sys.argv[1]):
                print("Path {0:s} not exist".format(os.path.basename(sys.argv[1])))
                sys.exit(1)
        if not os.path.exists(sys.argv[2]):
                print("Path {0:s} not exist".format(os.path.basename(sys.argv[2])))
                sys.exit(1)
	basemap = generateTextFile(sys.argv[1])
	postmap = generateTextFile(sys.argv[2])
	dcmp=dircmp(sys.argv[1], sys.argv[2])
	files=collectDiffFiles(dcmp)
	
	trueDiffs = sortAndCompare(files["differences"])

	for file in (files["baseline missing"]):
		print("{0:s} found in {1:s} but not {2:s}".format(file, os.path.basename(sys.argv[2]), os.path.basename(sys.argv[1])))
	for file in (files["post missing"]):
		print("{0:s} found in {1:s} but not {2:s}".format(file, os.path.basename(sys.argv[1]), os.path.basename(sys.argv[2])))
	diffs=Set()
	for file in trueDiffs:
		if os.path.splitext(file)[1].endswith('xls') or os.path.splitext(file)[1].endswith('xlsx'):
			for origin, mapping in basemap:
				if file in mapping:
					diffs.add(origin)
		else:
			diffs.add(file)
	for diff in diffs:
		print("{0:s} differs".format(diff))
	
if __name__ == '__main__':
    main()
